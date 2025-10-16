#!/usr/bin/env python3
"""
创建完整的模型对比表格和报告
包含所有重要参数和性能指标
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
import json
import yaml
from typing import Dict, List, Any, Optional
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.offline as pyo

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False
sns.set_style("whitegrid")

class ComprehensiveModelComparison:
    """综合模型对比分析器"""
    
    def __init__(self, batch_results_dir: str):
        self.batch_results_dir = Path(batch_results_dir)
        self.output_dir = self.batch_results_dir / "comprehensive_analysis"
        self.output_dir.mkdir(exist_ok=True)
        
        # 加载现有的CSV数据
        self.csv_file = self.batch_results_dir / "analysis" / "model_ranking.csv"
        if not self.csv_file.exists():
            raise FileNotFoundError(f"未找到模型排名CSV文件: {self.csv_file}")
        
        self.df = pd.read_csv(self.csv_file)
        logger.info(f"加载了 {len(self.df)} 个模型的数据")
        
    def enhance_model_data(self) -> pd.DataFrame:
        """增强模型数据，添加更多参数信息"""
        enhanced_df = self.df.copy()
        
        # 添加模型类别
        model_categories = {
            'FNO2D': 'Fourier Neural Operator',
            'SwinUNet': 'Vision Transformer',
            'UNet': 'Convolutional Network',
            'MLP': 'Multi-Layer Perceptron',
            'MLP_Mixer': 'MLP-based Architecture',
            'Hybrid': 'Hybrid Architecture',
            'UFNO_UNet': 'Hybrid FNO-UNet',
            'UNetPlusPlus': 'Enhanced UNet'
        }
        
        enhanced_df['模型类别'] = enhanced_df['模型'].map(model_categories)
        
        # 添加性能等级
        def get_performance_grade(rel_l2):
            if rel_l2 <= 0.02:
                return 'A+ (优秀)'
            elif rel_l2 <= 0.04:
                return 'A (良好)'
            elif rel_l2 <= 0.08:
                return 'B (中等)'
            elif rel_l2 <= 0.15:
                return 'C (一般)'
            else:
                return 'D (较差)'
        
        enhanced_df['性能等级'] = enhanced_df['Rel-L2'].apply(get_performance_grade)
        
        # 添加效率等级
        def get_efficiency_grade(time_minutes):
            if time_minutes <= 1:
                return 'A+ (极快)'
            elif time_minutes <= 3:
                return 'A (快速)'
            elif time_minutes <= 10:
                return 'B (中等)'
            elif time_minutes <= 30:
                return 'C (较慢)'
            else:
                return 'D (很慢)'
        
        enhanced_df['效率等级'] = enhanced_df['训练时间(分钟)'].apply(get_efficiency_grade)
        
        # 添加综合评分 (基于多个指标的加权平均)
        def calculate_composite_score(row):
            # 归一化各项指标 (越小越好的指标需要取倒数)
            rel_l2_norm = 1 / (1 + row['Rel-L2'])  # 越小越好
            mae_norm = 1 / (1 + row['MAE'])        # 越小越好
            psnr_norm = row['PSNR'] / 50           # 越大越好
            ssim_norm = row['SSIM']                # 越大越好
            time_norm = 1 / (1 + row['训练时间(分钟)'] / 10)  # 越小越好
            
            # 加权平均 (可以调整权重)
            weights = [0.3, 0.2, 0.2, 0.2, 0.1]  # rel_l2, mae, psnr, ssim, time
            score = (rel_l2_norm * weights[0] + 
                    mae_norm * weights[1] + 
                    psnr_norm * weights[2] + 
                    ssim_norm * weights[3] + 
                    time_norm * weights[4])
            return score * 100  # 转换为0-100分
        
        enhanced_df['综合评分'] = enhanced_df.apply(calculate_composite_score, axis=1)
        
        # 重新排序列
        column_order = [
            '模型', '模型类别', '性能等级', '效率等级', '综合评分',
            'Rel-L2', 'MAE', 'PSNR', 'SSIM', '最佳验证损失',
            '训练时间(分钟)', '参数量(M)', 'BRMSE', 'CRMSE'
        ]
        
        enhanced_df = enhanced_df[column_order]
        
        # 按综合评分排序
        enhanced_df = enhanced_df.sort_values('综合评分', ascending=False).reset_index(drop=True)
        enhanced_df.index = enhanced_df.index + 1  # 从1开始排名
        
        return enhanced_df
    
    def create_excel_report(self, enhanced_df: pd.DataFrame) -> Path:
        """创建Excel格式的详细报告"""
        excel_file = self.output_dir / "comprehensive_model_comparison.xlsx"
        
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # 主要对比表
            enhanced_df.to_excel(writer, sheet_name='模型对比', index_label='排名')
            
            # 性能分析表
            performance_df = enhanced_df[['模型', 'Rel-L2', 'MAE', 'PSNR', 'SSIM', '综合评分']].copy()
            performance_df.to_excel(writer, sheet_name='性能分析', index=False)
            
            # 效率分析表
            efficiency_df = enhanced_df[['模型', '训练时间(分钟)', '参数量(M)', '效率等级']].copy()
            efficiency_df.to_excel(writer, sheet_name='效率分析', index=False)
            
            # 统计汇总
            stats_data = {
                '指标': ['平均值', '标准差', '最小值', '最大值', '中位数'],
                'Rel-L2': [
                    enhanced_df['Rel-L2'].mean(),
                    enhanced_df['Rel-L2'].std(),
                    enhanced_df['Rel-L2'].min(),
                    enhanced_df['Rel-L2'].max(),
                    enhanced_df['Rel-L2'].median()
                ],
                'PSNR': [
                    enhanced_df['PSNR'].mean(),
                    enhanced_df['PSNR'].std(),
                    enhanced_df['PSNR'].min(),
                    enhanced_df['PSNR'].max(),
                    enhanced_df['PSNR'].median()
                ],
                'SSIM': [
                    enhanced_df['SSIM'].mean(),
                    enhanced_df['SSIM'].std(),
                    enhanced_df['SSIM'].min(),
                    enhanced_df['SSIM'].max(),
                    enhanced_df['SSIM'].median()
                ],
                '训练时间(分钟)': [
                    enhanced_df['训练时间(分钟)'].mean(),
                    enhanced_df['训练时间(分钟)'].std(),
                    enhanced_df['训练时间(分钟)'].min(),
                    enhanced_df['训练时间(分钟)'].max(),
                    enhanced_df['训练时间(分钟)'].median()
                ]
            }
            stats_df = pd.DataFrame(stats_data)
            stats_df.to_excel(writer, sheet_name='统计汇总', index=False)
        
        # 美化Excel格式
        self._format_excel(excel_file)
        
        logger.info(f"Excel报告已保存: {excel_file}")
        return excel_file
    
    def _format_excel(self, excel_file: Path):
        """美化Excel格式"""
        wb = openpyxl.load_workbook(excel_file)
        
        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 格式化每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 设置标题行样式
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # 设置数据行样式
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = center_alignment
                    cell.border = border
            
            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_file)
    
    def create_markdown_table(self, enhanced_df: pd.DataFrame) -> Path:
        """创建Markdown格式的对比表"""
        md_file = self.output_dir / "model_comparison_table.md"
        
        with open(md_file, 'w', encoding='utf-8') as f:
            f.write("# 🚀 Sparse2Full 模型性能对比表\n\n")
            f.write(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            # 主要对比表
            f.write("## 📊 综合性能对比\n\n")
            
            # 创建简化的表格用于Markdown显示
            display_df = enhanced_df[['模型', '模型类别', '性能等级', '效率等级', 
                                    'Rel-L2', 'PSNR', 'SSIM', '训练时间(分钟)', '综合评分']].copy()
            
            # 格式化数值
            display_df['Rel-L2'] = display_df['Rel-L2'].apply(lambda x: f"{x:.4f}")
            display_df['PSNR'] = display_df['PSNR'].apply(lambda x: f"{x:.2f}")
            display_df['SSIM'] = display_df['SSIM'].apply(lambda x: f"{x:.4f}")
            display_df['训练时间(分钟)'] = display_df['训练时间(分钟)'].apply(lambda x: f"{x:.2f}")
            display_df['综合评分'] = display_df['综合评分'].apply(lambda x: f"{x:.1f}")
            
            f.write(display_df.to_markdown(index=True))
            f.write("\n\n")
            
            # 性能分析
            f.write("## 🏆 性能分析\n\n")
            f.write("### 最佳模型 (Top 3)\n\n")
            top3 = enhanced_df.head(3)
            for i, (_, row) in enumerate(top3.iterrows(), 1):
                medal = ["🥇", "🥈", "🥉"][i-1]
                f.write(f"{medal} **{row['模型']}** ({row['模型类别']})\n")
                f.write(f"   - Rel-L2: {row['Rel-L2']:.4f}\n")
                f.write(f"   - PSNR: {row['PSNR']:.2f} dB\n")
                f.write(f"   - SSIM: {row['SSIM']:.4f}\n")
                f.write(f"   - 训练时间: {row['训练时间(分钟)']:.2f} 分钟\n")
                f.write(f"   - 综合评分: {row['综合评分']:.1f}/100\n\n")
            
            # 效率分析
            f.write("## ⚡ 效率分析\n\n")
            fastest_models = enhanced_df.nsmallest(3, '训练时间(分钟)')
            f.write("### 训练速度最快 (Top 3)\n\n")
            for _, row in fastest_models.iterrows():
                f.write(f"- **{row['模型']}**: {row['训练时间(分钟)']:.2f} 分钟 ({row['效率等级']})\n")
            f.write("\n")
            
            # 统计信息
            f.write("## 📈 统计信息\n\n")
            f.write(f"- **总模型数**: {len(enhanced_df)}\n")
            f.write(f"- **平均Rel-L2**: {enhanced_df['Rel-L2'].mean():.4f} ± {enhanced_df['Rel-L2'].std():.4f}\n")
            f.write(f"- **平均PSNR**: {enhanced_df['PSNR'].mean():.2f} ± {enhanced_df['PSNR'].std():.2f} dB\n")
            f.write(f"- **平均SSIM**: {enhanced_df['SSIM'].mean():.4f} ± {enhanced_df['SSIM'].std():.4f}\n")
            f.write(f"- **平均训练时间**: {enhanced_df['训练时间(分钟)'].mean():.2f} ± {enhanced_df['训练时间(分钟)'].std():.2f} 分钟\n\n")
            
            # 使用建议
            f.write("## 💡 使用建议\n\n")
            best_model = enhanced_df.iloc[0]
            fastest_model = enhanced_df.loc[enhanced_df['训练时间(分钟)'].idxmin()]
            
            f.write(f"### 🎯 最佳性能推荐\n")
            f.write(f"**{best_model['模型']}** - 综合评分最高 ({best_model['综合评分']:.1f}/100)\n")
            f.write(f"- 适用场景: 对精度要求极高的应用\n")
            f.write(f"- 优势: {best_model['性能等级']}\n\n")
            
            f.write(f"### ⚡ 最佳效率推荐\n")
            f.write(f"**{fastest_model['模型']}** - 训练时间最短 ({fastest_model['训练时间(分钟)']:.2f} 分钟)\n")
            f.write(f"- 适用场景: 快速原型开发和实时应用\n")
            f.write(f"- 优势: {fastest_model['效率等级']}\n\n")
        
        logger.info(f"Markdown表格已保存: {md_file}")
        return md_file
    
    def create_interactive_charts(self, enhanced_df: pd.DataFrame) -> List[Path]:
        """创建交互式图表"""
        chart_files = []
        
        # 1. 性能对比雷达图
        radar_file = self.output_dir / "performance_radar_chart.html"
        self._create_radar_chart(enhanced_df, radar_file)
        chart_files.append(radar_file)
        
        # 2. 散点图矩阵
        scatter_file = self.output_dir / "performance_scatter_matrix.html"
        self._create_scatter_matrix(enhanced_df, scatter_file)
        chart_files.append(scatter_file)
        
        # 3. 综合评分条形图
        bar_file = self.output_dir / "composite_score_chart.html"
        self._create_composite_score_chart(enhanced_df, bar_file)
        chart_files.append(bar_file)
        
        return chart_files
    
    def _create_radar_chart(self, df: pd.DataFrame, output_file: Path):
        """创建雷达图"""
        # 选择前5个模型
        top5_df = df.head(5)
        
        # 归一化指标 (0-1范围)
        metrics = ['Rel-L2', 'MAE', 'PSNR', 'SSIM', '训练时间(分钟)']
        normalized_data = {}
        
        for metric in metrics:
            if metric in ['Rel-L2', 'MAE', '训练时间(分钟)']:
                # 越小越好的指标，取倒数后归一化
                values = 1 / (1 + top5_df[metric])
            else:
                # 越大越好的指标，直接归一化
                values = top5_df[metric] / top5_df[metric].max()
            normalized_data[metric] = values
        
        fig = go.Figure()
        
        for i, (_, row) in enumerate(top5_df.iterrows()):
            values = [normalized_data[metric].iloc[i] for metric in metrics]
            values.append(values[0])  # 闭合雷达图
            
            fig.add_trace(go.Scatterpolar(
                r=values,
                theta=metrics + [metrics[0]],
                fill='toself',
                name=row['模型'],
                line_color=px.colors.qualitative.Set1[i]
            ))
        
        fig.update_layout(
            polar=dict(
                radialaxis=dict(
                    visible=True,
                    range=[0, 1]
                )),
            showlegend=True,
            title="模型性能雷达图 (Top 5)",
            font=dict(size=14)
        )
        
        pyo.plot(fig, filename=str(output_file), auto_open=False)
    
    def _create_scatter_matrix(self, df: pd.DataFrame, output_file: Path):
        """创建散点图矩阵"""
        metrics = ['Rel-L2', 'PSNR', 'SSIM', '训练时间(分钟)']
        
        fig = make_subplots(
            rows=len(metrics), cols=len(metrics),
            subplot_titles=[f"{m1} vs {m2}" for m1 in metrics for m2 in metrics]
        )
        
        for i, metric1 in enumerate(metrics):
            for j, metric2 in enumerate(metrics):
                if i == j:
                    # 对角线显示直方图
                    fig.add_trace(
                        go.Histogram(x=df[metric1], name=metric1, showlegend=False),
                        row=i+1, col=j+1
                    )
                else:
                    # 非对角线显示散点图
                    fig.add_trace(
                        go.Scatter(
                            x=df[metric2], y=df[metric1],
                            mode='markers+text',
                            text=df['模型'],
                            textposition="top center",
                            name=f"{metric1} vs {metric2}",
                            showlegend=False
                        ),
                        row=i+1, col=j+1
                    )
        
        fig.update_layout(
            height=800, width=800,
            title_text="模型性能指标散点图矩阵"
        )
        
        pyo.plot(fig, filename=str(output_file), auto_open=False)
    
    def _create_composite_score_chart(self, df: pd.DataFrame, output_file: Path):
        """创建综合评分图表"""
        fig = go.Figure()
        
        # 添加条形图
        fig.add_trace(go.Bar(
            x=df['模型'],
            y=df['综合评分'],
            text=[f"{score:.1f}" for score in df['综合评分']],
            textposition='auto',
            marker_color=px.colors.qualitative.Set3,
            name='综合评分'
        ))
        
        fig.update_layout(
            title="模型综合评分对比",
            xaxis_title="模型",
            yaxis_title="综合评分 (0-100)",
            showlegend=False,
            height=500
        )
        
        pyo.plot(fig, filename=str(output_file), auto_open=False)
    
    def generate_comprehensive_report(self) -> Dict[str, Path]:
        """生成综合报告"""
        logger.info("开始生成综合模型对比报告...")
        
        # 1. 增强数据
        enhanced_df = self.enhance_model_data()
        
        # 2. 保存增强后的CSV
        enhanced_csv = self.output_dir / "enhanced_model_comparison.csv"
        enhanced_df.to_csv(enhanced_csv, index_label='排名', encoding='utf-8-sig')
        
        # 3. 创建Excel报告
        excel_file = self.create_excel_report(enhanced_df)
        
        # 4. 创建Markdown表格
        markdown_file = self.create_markdown_table(enhanced_df)
        
        # 5. 创建交互式图表
        chart_files = self.create_interactive_charts(enhanced_df)
        
        # 6. 生成汇总信息
        summary = {
            'total_models': len(enhanced_df),
            'best_model': enhanced_df.iloc[0]['模型'],
            'fastest_model': enhanced_df.loc[enhanced_df['训练时间(分钟)'].idxmin(), '模型'],
            'avg_rel_l2': enhanced_df['Rel-L2'].mean(),
            'avg_psnr': enhanced_df['PSNR'].mean(),
            'avg_ssim': enhanced_df['SSIM'].mean(),
            'avg_training_time': enhanced_df['训练时间(分钟)'].mean()
        }
        
        summary_file = self.output_dir / "comparison_summary.json"
        with open(summary_file, 'w', encoding='utf-8') as f:
            json.dump(summary, f, indent=2, ensure_ascii=False)
        
        results = {
            'enhanced_csv': enhanced_csv,
            'excel_report': excel_file,
            'markdown_table': markdown_file,
            'summary': summary_file
        }
        
        # 添加图表文件
        for i, chart_file in enumerate(chart_files):
            results[f'chart_{i+1}'] = chart_file
        
        logger.info(f"综合报告生成完成，输出目录: {self.output_dir}")
        return results


def main():
    """主函数"""
    batch_results_dir = "f:/Zhaoyang/Sparse2Full/runs/batch_retrain_20251015_032934"
    
    try:
        # 创建对比分析器
        comparator = ComprehensiveModelComparison(batch_results_dir)
        
        # 生成综合报告
        results = comparator.generate_comprehensive_report()
        
        print("\n🎉 综合模型对比报告生成完成！")
        print(f"📁 输出目录: {comparator.output_dir}")
        print("\n生成的文件:")
        for name, path in results.items():
            print(f"  - {name}: {path.name}")
        
        return results
        
    except Exception as e:
        logger.error(f"生成报告时出错: {e}")
        raise


if __name__ == "__main__":
    main()